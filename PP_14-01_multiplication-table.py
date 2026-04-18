# Create a program multiplicationTable.py that takes a number N from the command line and creates an NxN multiplicaiton table in an excel spreadsheet.

import openpyxl, sys
from openpyxl.styles import Font
# Check if the user provided a value 
if len(sys.argv) != 2:
    print(f"{sys.argv[0]}: Creates an NxN multiplicaiton table. Usage: {sys.argv[0]} <number>")
    sys.exit(1)
try:
    n = int(sys.argv[1])
except ValueError:
    print(f"Error: {sys.argv[1]} is not an integer")
    sys.exit(1)
if n <= 0:
    print("Error: number must be positive")
    sys.exit(1)

wb = openpyxl.Workbook()
sheet = wb.active
f_bold = Font(bold=True)

# Set up bold labels
for i in range(2, n + 2):
    sheet.cell(row = 1, column = i).value = i - 1
    sheet.cell(row = 1, column = i).font = f_bold
    sheet.cell(column = 1, row = i).value = i -1
    sheet.cell(column = 1, row = i).font = f_bold

for i in range(1, n + 1):
    for j in range(1, n + 1):
        m_value = i * j
        sheet.cell(row = i+1, column = j+1).value = m_value
wb.save('multiplicationTable.xlsx')



