# Excel-to-CSV Converter 
# Excel can save a spreadsheet to a CSV file with a few mouse clicks, but if you had to convert hundered of Excel files to CSVs, it would take hours of clicking.
# Using the openpyxl module fro Chapter 14, write a program that reads all the Excel files in the current working directory and outputs them as CSV files.
# A single Excel file might contain multiplel sheets; you'll have to create one CSV file per sheet. The filenames of the CSV files should be <excelfilename>_<sheet title>.csv, where
# <excel filename> is the filename of the Excel file without the file extension (for example, spam_data, not spam_data.xlsx) and <sheet title> is the string from the
# WorkSheet object's title variable.

import openpyxl
import csv
import os

for excel_file in os.listdir('.'):
    # Skip non-xlsx files, load the workbook object.
    if not excel_file.endswith(".xlsx"):
        continue

    try:
        wb = openpyxl.load_workbook(excel_file)
    except Exception as e:
        print(f"Error: {e} with file {excel_file}")
        continue

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        # Loop through every sheet in the workbook.
        # Create the CSV filename from the Excel filename and sheet title.
        # Create the csv.writer object for this CSV file.
        csv_name = f"{excel_file[:-5]}_{sheet_name}.csv"

        with open(csv_name, 'w', newline='', encoding='utf-8') as output_file:
            output_writer = csv.writer(output_file)

            # Loop through every row in the sheet.
            for row_num in range(1, sheet.max_row + 1):
                row_data = []   # Append each cell to this list.

                # Loop through each cell in the row.
                for col_num in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row=row_num, column=col_num).value
                    # Append each cell's data to row_data
                    row_data.append(cell_value)

                # Write the row_data list to the CSV file.
                output_writer.writerow(row_data)
        
        print(f"Created {csv_name}")

