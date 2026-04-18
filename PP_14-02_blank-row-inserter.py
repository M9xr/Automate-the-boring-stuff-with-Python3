# Create a progrma blankRowInserter.py that takes two integers and a filename string as command line arguments. Let's call the first integer N
# and the second integer M. Starting at row N, the program should insert M blank rows into the spreadsheet.
# You can wrtie this program by reading in the contents of the spreadsheet. Then, when writing out the new spreadsheet, use a for loop
# to copy the first N lines. For the remaing lines, add M to the row number in the output spreadsheet.

import openpyxl, sys

# Check if the user provided a value
if len(sys.argv) != 4:
    print(f"{sys.argv[0]}: Inserts M blank rows, starting at row N. Usage: {sys.argv[0]} <N> <M> <file_name.xlsx>")
    sys.exit(1)

N = int(sys.argv[1])
M = int(sys.argv[2])
file_name = sys.argv[3]

wb = openpyxl.load_workbook(file_name)
sheet = wb.active

lines_to_copy = []
for i in range(1, sheet.max_row + 1):
    line_to_copy = []
    for j in range(1, sheet.max_column + 1):
        line_to_copy.append(sheet.cell(row = i, column = j).value)
    lines_to_copy.append(line_to_copy)


wb2 = openpyxl.Workbook()
sheet2 = wb2.active
for i in range(1, N):
    for j in range(1, sheet.max_column + 1):
        sheet2.cell(row = i, column = j).value = lines_to_copy[i-1][j-1]

for i in range(N, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        sheet2.cell(row = i+M, column = j).value = lines_to_copy[i-1][j-1]
wb2.save('output3.xlsx')



